from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .services import EventService

HEADERS = ["행사", "대분류", "중분류", "항목", "세부내용", "상태", "수량", "단위",
           "작업 시작일", "마감일", "담당자(PM)", "업체", "업체담당자", "업체담당자 전화번호", "VAT 구분", "메모"]


def _task_rows(db, event_id=None):
    sql = """SELECT e.name event_name,t.major,t.minor,t.name,t.detail,t.status,t.quantity,t.unit,
             t.planned_start,t.due_date,pm.name pm_assignee,v.name vendor,p.name vendor_assignee,p.phone,
             t.vat_type,t.note
             FROM event_tasks t JOIN events e ON e.id=t.event_id
             LEFT JOIN contacts pm ON pm.id=t.pm_assignee_id
             LEFT JOIN contacts p ON p.id=t.assignee_id LEFT JOIN contacts v ON v.id=t.vendor_id WHERE t.is_removed=0"""
    params = []
    if event_id is not None: sql += " AND t.event_id=?"; params.append(event_id)
    sql += " ORDER BY e.start_date,t.due_date,t.sort_order"
    return db.query(sql, params)


def export_csv(db, destination: Path, event_id=None):
    destination = Path(destination); destination.parent.mkdir(parents=True, exist_ok=True)
    with destination.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle); writer.writerow(HEADERS)
        for row in _task_rows(db, event_id): writer.writerow(list(row))
    return destination


def _style_sheet(ws, widths):
    ws.freeze_panes = "A2"; ws.auto_filter.ref = ws.dimensions
    fill = PatternFill("solid", fgColor="F25B24")
    for cell in ws[1]:
        cell.fill = fill; cell.font = Font(bold=True, color="FFFFFF"); cell.alignment = Alignment(horizontal="center")
    for index, width in enumerate(widths, 1): ws.column_dimensions[get_column_letter(index)].width = width


def export_excel(db, destination: Path, event_id=None):
    destination = Path(destination); destination.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook(); ws = wb.active; ws.title = "체크리스트"; ws.append(HEADERS)
    for row in _task_rows(db, event_id): ws.append(list(row))
    _style_sheet(ws, [24,10,14,22,36,12,10,10,14,14,16,18,18,18,12,36])

    service = EventService(db)
    events = service.list_events() if event_id is None else [service.get_event(event_id)]
    detail = wb.create_sheet("상세 정산")
    detail_headers = ["행사", "대분류", "중분류", "항목", "수량", "단위", "행사 단가", "공급가", "VAT 구분", "VAT", "합계", "업체", "메모"]
    detail.append(detail_headers)
    summary_ws = wb.create_sheet("정산 요약")
    summary_ws.append(["행사", "구분", "공급가", "VAT", "VAT 포함 합계", "입력 예산", "예산 기준", "차이"])
    for event in events:
        summary = service.settlement_summary(int(event["id"]))
        for item in summary["items"]:
            detail.append([event["name"], item["major"], item["minor"], item["name"], item["quantity"], item["unit"],
                           item["unit_price"], item["supply"], "VAT 10%" if item["vat_type"] == "TAXABLE" else "면세",
                           item["vat"], item["total"], item["vendor_name"], item["note"]])
        for major, subtotal in summary["categories"].items():
            summary_ws.append([event["name"], f"{major} 소계", subtotal["supply"], subtotal["vat"], subtotal["total"], "", "", ""])
        mode = {"INCLUDED": "VAT 포함", "EXCLUDED": "VAT 별도", "UNSET": "미선택"}[summary["budget_tax_mode"]]
        summary_ws.append([event["name"], "전체 합계", summary["supply"], summary["vat"], summary["total"],
                           summary["budget"], mode, summary["difference"]])
    _style_sheet(detail, [24,10,14,24,10,9,14,14,12,12,14,18,30])
    _style_sheet(summary_ws, [24,18,15,13,17,15,13,15])

    events_ws = wb.create_sheet("행사")
    events_ws.append(["ID", "행사명", "준비 시작일", "최종 행사일", "장소", "주최/주관", "예산", "예산 VAT 기준", "PM 업체"])
    for row in db.query("""SELECT e.id,e.name,e.start_date,e.end_date,e.location,e.organizer,e.budget,e.budget_tax_mode,v.name
                           FROM events e LEFT JOIN contacts v ON v.id=e.pm_vendor_id ORDER BY e.start_date"""):
        events_ws.append(list(row))
    contacts = wb.create_sheet("업체·담당자")
    contacts.append(["구분", "이름", "소속 업체", "연락처", "역할/분야"])
    for row in db.query("""SELECT c.kind,c.name,company.name company,c.phone,c.role_note FROM contacts c
                           LEFT JOIN contacts company ON company.id=c.company_id ORDER BY c.kind,company.name,c.name"""):
        contacts.append(list(row))
    wb.save(destination); load_workbook(destination, read_only=True).close(); return destination
