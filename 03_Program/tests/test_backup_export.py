from __future__ import annotations

from datetime import date

from openpyxl import load_workbook

from event_checklist.backup import create_backup, create_manual_backup, create_rotating_auto_backup, restore_backup
from event_checklist.export import export_csv, export_excel
from event_checklist.services import EventService


def _event(db):
    service = EventService(db)
    ids = [row["id"] for row in db.query("SELECT id FROM master_items ORDER BY id LIMIT 2")]
    return service.create_event("내보내기 행사", date(2026, 9, 1), None, ids)


def test_backup_and_restore(db, tmp_path):
    _event(db)
    backup = create_backup(db, tmp_path / "backup.db")
    db.execute("DELETE FROM events")
    assert db.one("SELECT COUNT(*) count FROM events")["count"] == 0
    restore_backup(db, backup)
    assert db.one("SELECT COUNT(*) count FROM events")["count"] == 1
    assert db.one("SELECT COUNT(*) count FROM event_tasks")["count"] == 2


def test_excel_and_csv_export(db, tmp_path):
    event_id = _event(db)
    xlsx = export_excel(db, tmp_path / "output.xlsx", event_id)
    csv_file = export_csv(db, tmp_path / "output.csv", event_id)
    assert xlsx.exists() and csv_file.exists()
    workbook = load_workbook(xlsx, read_only=True)
    assert workbook["체크리스트"].max_row == 3
    assert workbook["체크리스트"]["A2"].value == "내보내기 행사"
    assert {"상세 정산", "정산 요약"} <= set(workbook.sheetnames)
    assert workbook["상세 정산"]["A2"].value == "내보내기 행사"
    assert workbook["정산 요약"].max_row >= 2
    workbook.close()
    assert "내보내기 행사" in csv_file.read_text(encoding="utf-8-sig")


def test_auto_backup_rotation_keeps_ten_and_preserves_manual(db, tmp_path):
    backup_directory = tmp_path / "backups"
    manual = create_manual_backup(db, backup_directory)
    for index in range(12):
        db.execute("UPDATE contacts SET phone=? WHERE id=1", (str(index),))
        create_rotating_auto_backup(db, backup_directory, keep=10)
    assert manual.exists()
    assert len(list(backup_directory.glob("auto_event_flow_*.db"))) == 10
